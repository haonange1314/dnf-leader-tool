from io import BytesIO

from openpyxl import load_workbook
from PIL import Image

from app.application.schedule_exports import snapshot_png, snapshot_text, snapshot_workbook


def _snapshot() -> dict[str, object]:
    return {
        "name": "周六团",
        "participants": [
            {
                "id": "participant-1",
                "playerNameSnapshot": "玩家一",
                "characterNameSnapshot": "角色一",
                "professionSnapshot": "剑魂",
                "roleTypeSnapshot": "DAMAGE",
                "damageScoreSnapshot": "500.00",
                "bufferScoreSnapshot": None,
                "isSelected": True,
                "unassignedReason": None,
            },
            {
                "id": "participant-2",
                "playerNameSnapshot": "玩家二",
                "characterNameSnapshot": "角色二",
                "professionSnapshot": "奶妈",
                "roleTypeSnapshot": "BUFFER",
                "damageScoreSnapshot": None,
                "bufferScoreSnapshot": "50.0",
                "isSelected": True,
                "unassignedReason": {"code": "UNASSIGNED_CAPACITY"},
            },
        ],
        "waves": [
            {
                "waveNo": 1,
                "damageTotal": "500.00",
                "bufferTotal": "0.0",
                "specialAssignments": [
                    {"participantId": "participant-1", "ruleCode": "TREASURE_CORE"}
                ],
                "teams": [
                    {
                        "displayNameSnapshot": "红队",
                        "compositionCode": "INCOMPLETE",
                        "damageTotal": "500.00",
                        "bufferTotal": "0.0",
                        "slots": [
                            {"slotNo": 1, "participantId": "participant-1"},
                            {"slotNo": 2, "participantId": None},
                        ],
                    }
                ],
            }
        ],
        "issues": [
            {
                "severity": "ERROR",
                "code": "TEAM_INCOMPLETE",
                "message_params": {"waveNo": 1, "teamKey": "RED"},
            }
        ],
    }


def test_draft_exports_include_visible_watermarks() -> None:
    snapshot = _snapshot()

    text = snapshot_text(snapshot, "revision 3", draft=True)
    png = snapshot_png(snapshot, "revision 3", draft=True)

    assert text.startswith("【草稿】周六团 · revision 3")
    assert "玩家一·角色一【核心】" in text
    assert png.read(8) == b"\x89PNG\r\n\x1a\n"
    png.seek(0)
    with Image.open(png) as image:
        assert image.format == "PNG"
        assert image.width >= 1200
        assert image.height >= 240


def test_excel_export_contains_all_review_sheets() -> None:
    output = snapshot_workbook(_snapshot(), "发布版本 v1")
    workbook = load_workbook(BytesIO(output.read()), read_only=True)

    assert workbook.sheetnames == ["排表总览", "未分配", "强度统计", "问题清单"]
    assert workbook["排表总览"]["A3"].value == 1
    assert workbook["未分配"]["A3"].value == "玩家二"
    assert workbook["问题清单"]["B3"].value == "TEAM_INCOMPLETE"
