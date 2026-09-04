from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.imports.characters import (
    HEADERS,
    CharacterExportRow,
    CharacterImportDefaults,
    build_roster_workbook,
    build_template,
    parse_character_workbook,
)


def test_template_can_be_parsed() -> None:
    content = build_template()
    rows = parse_character_workbook(content, 100)
    workbook = load_workbook(BytesIO(content), read_only=True)
    sheet = workbook["角色数据"]

    assert tuple(cell.value for cell in sheet[1]) == HEADERS
    assert "角色名" not in HEADERS
    assert len(rows) == 1
    assert rows[0].payload["role_type"] == "DAMAGE"
    assert rows[0].payload["is_treasure_damage"] is False
    assert rows[0].payload["is_fixed_lead_team_buffer"] is False
    assert rows[0].payload["is_group_hunt"] is False
    assert rows[0].payload["default_raid_participant"] is True
    assert not rows[0].errors


def test_parser_reports_duplicate_and_invalid_row() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "角色数据"
    sheet.append(HEADERS)
    sheet.append((1, "玩家A", "剑魂", "C", "120亿", "是", "", "", "是"))
    sheet.append((2, "玩家A", "剑魂", "未知", "x", "?", "", "", "是"))
    stream = BytesIO()
    workbook.save(stream)

    rows = parse_character_workbook(stream.getvalue(), 100)

    assert not rows[0].errors
    assert {error["code"] for error in rows[1].errors} >= {
        "DUPLICATE_ROW",
        "INVALID_ROLE",
        "INVALID_SCORE",
        "INVALID_BOOLEAN",
    }


def test_parser_accepts_latest_roster_layout_and_two_decimal_buffer_score() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "米歇尔统计"
    sheet.append(HEADERS)
    sheet.append((1, "剑来", "奶萝", "奶", "4.75", "", "是", "", ""))
    sheet.append((2, "剑来", "光兵", "C", 1500, "", "", "是", "是"))
    workbook.create_sheet("军团-剑来小队").append(("这不是导入页",))
    stream = BytesIO()
    workbook.save(stream)

    rows = parse_character_workbook(stream.getvalue(), 100)

    assert len(rows) == 2
    assert rows[0].payload["buffer_score"] == "4.75"
    assert rows[0].payload["is_fixed_lead_team_buffer"] is True
    assert rows[0].payload["default_raid_participant"] is True
    assert rows[1].payload["damage_score"] == "1500"
    assert rows[1].payload["is_group_hunt"] is True
    assert rows[1].payload["default_raid_participant"] is True
    assert not rows[0].errors
    assert not rows[1].errors


def test_parser_keeps_legacy_seven_column_template_compatible() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "角色数据"
    sheet.append(("玩家称呼", "职业", "类型", "伤害/增益量", "秘宝C", "默认参团", "备注"))
    sheet.append(("旧玩家", "剑魂", "C", 120, "否", "是", "旧模板"))
    stream = BytesIO()
    workbook.save(stream)

    rows = parse_character_workbook(stream.getvalue(), 100)

    assert len(rows) == 1
    assert rows[0].payload["note"] == "旧模板"
    assert rows[0].payload["is_fixed_lead_team_buffer"] is False
    assert rows[0].payload["is_group_hunt"] is False
    assert not rows[0].errors


def test_parser_accepts_new_schedule_default_selection_header() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "角色数据"
    sheet.append(
        ("玩家昵称", "职业", "类型", "模拟伤害亿/增益量万", "默认加入新排表")
    )
    sheet.append(("玩家A", "剑魂", "C", 120, "否"))
    stream = BytesIO()
    workbook.save(stream)

    rows = parse_character_workbook(stream.getvalue(), 100)

    assert len(rows) == 1
    assert rows[0].payload["default_raid_participant"] is False
    assert not rows[0].errors


def test_blank_boolean_columns_use_configurable_defaults() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "角色数据"
    sheet.append(HEADERS)
    sheet.append((1, "玩家A", "剑魂", "C", 120, "", "", "", ""))
    stream = BytesIO()
    workbook.save(stream)

    rows = parse_character_workbook(stream.getvalue(), 100)

    assert rows[0].payload["is_treasure_damage"] is False
    assert rows[0].payload["is_fixed_lead_team_buffer"] is False
    assert rows[0].payload["is_group_hunt"] is False
    assert rows[0].payload["default_raid_participant"] is True

    configured_rows = parse_character_workbook(
        stream.getvalue(),
        100,
        CharacterImportDefaults(default_raid_participant=False),
    )
    assert configured_rows[0].payload["default_raid_participant"] is False


def test_template_explains_configured_blank_defaults() -> None:
    content = build_template(CharacterImportDefaults(default_raid_participant=False))
    workbook = load_workbook(BytesIO(content), read_only=True)
    notes = {
        row[0].value: row[1].value
        for row in workbook["填写说明"].iter_rows(min_row=2, max_col=2)
    }

    assert "空白按否处理" in notes["是否秘宝C"]
    assert "空白按否处理" in notes["固定红队奶"]
    assert "空白按否处理" in notes["是否群猎"]
    assert "空白按否处理" in notes["是否参与团本"]


def test_parser_rejects_empty_roster_to_prevent_accidental_full_deactivation() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "角色数据"
    sheet.append(HEADERS)
    stream = BytesIO()
    workbook.save(stream)

    with pytest.raises(ValueError, match="没有角色数据"):
        parse_character_workbook(stream.getvalue(), 100)


def test_parser_rejects_fractional_damage_score() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "角色数据"
    sheet.append(HEADERS)
    sheet.append((1, "玩家A", "剑魂", "C", "120.5", "否", "否", "否", "是"))
    stream = BytesIO()
    workbook.save(stream)

    rows = parse_character_workbook(stream.getvalue(), 100)

    assert rows[0].errors == [
        {"code": "DAMAGE_SCORE_NOT_INTEGER", "message": "C 伤害必须为整数"}
    ]


def test_current_roster_export_round_trips_in_display_order() -> None:
    content = build_roster_workbook(
        [
            CharacterExportRow(
                player_name="玩家A",
                profession="剑魂",
                role_type="DAMAGE",
                damage_score=Decimal("120.00"),
                buffer_score=None,
                is_treasure_damage=True,
                is_fixed_lead_team_buffer=False,
                is_group_hunt=True,
                default_raid_participant=True,
            ),
            CharacterExportRow(
                player_name="玩家A",
                profession="奶妈",
                role_type="BUFFER",
                damage_score=None,
                buffer_score=Decimal("4.75"),
                is_treasure_damage=False,
                is_fixed_lead_team_buffer=True,
                is_group_hunt=False,
                default_raid_participant=False,
            ),
        ]
    )

    rows = parse_character_workbook(content, 100)

    assert [row.payload["profession"] for row in rows] == ["剑魂", "奶妈"]
    assert rows[0].payload["damage_score"] == "120"
    assert rows[0].payload["is_treasure_damage"] is True
    assert rows[0].payload["is_group_hunt"] is True
    assert rows[1].payload["buffer_score"] == "4.75"
    assert rows[1].payload["is_fixed_lead_team_buffer"] is True
    assert rows[1].payload["default_raid_participant"] is False
    assert not any(row.errors for row in rows)


def test_current_roster_export_rejects_legacy_fractional_damage() -> None:
    with pytest.raises(ValueError, match="仍有小数 C 伤害"):
        build_roster_workbook(
            [
                CharacterExportRow(
                    player_name="玩家A",
                    profession="剑魂",
                    role_type="DAMAGE",
                    damage_score=Decimal("120.5"),
                    buffer_score=None,
                    is_treasure_damage=False,
                    is_fixed_lead_team_buffer=False,
                    is_group_hunt=False,
                    default_raid_participant=True,
                )
            ]
        )
