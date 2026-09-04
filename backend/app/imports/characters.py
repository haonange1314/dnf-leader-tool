from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.personnel import normalize_key
from app.schemas.personnel import CharacterCreate, CharacterRole

HEADERS = (
    "序号",
    "玩家昵称",
    "职业",
    "类型",
    "模拟伤害亿/增益量万",
    "是否秘宝C",
    "固定红队奶",
    "是否群猎",
    "是否参与团本",
)

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "sequence": ("序号",),
    "player_name": ("玩家昵称", "玩家称呼"),
    "profession": ("职业",),
    "role_type": ("类型",),
    "score": ("模拟伤害亿/增益量万", "伤害/增益量"),
    "is_treasure_damage": ("是否秘宝C", "秘宝C"),
    "is_fixed_lead_team_buffer": ("固定红队奶",),
    "is_group_hunt": ("是否群猎",),
    "default_raid_participant": (
        "是否参与团本",
        "默认参团",
        "默认加入新排表",
    ),
    "note": ("备注",),
}
REQUIRED_FIELDS = ("player_name", "profession", "role_type", "score")


@dataclass(frozen=True)
class CharacterImportDefaults:
    is_treasure_damage: bool = False
    is_fixed_lead_team_buffer: bool = False
    is_group_hunt: bool = False
    default_raid_participant: bool = True


DEFAULT_CHARACTER_IMPORT_DEFAULTS = CharacterImportDefaults()


@dataclass(frozen=True)
class ParsedRow:
    row_no: int
    payload: dict[str, Any]
    errors: list[dict[str, str]]


@dataclass(frozen=True)
class CharacterExportRow:
    player_name: str
    profession: str
    role_type: str
    damage_score: Decimal | None
    buffer_score: Decimal | None
    is_treasure_damage: bool
    is_fixed_lead_team_buffer: bool
    is_group_hunt: bool
    default_raid_participant: bool


def build_template(
    defaults: CharacterImportDefaults = DEFAULT_CHARACTER_IMPORT_DEFAULTS,
) -> bytes:
    workbook = _base_workbook(defaults)
    sheet = workbook.active
    assert sheet is not None
    sheet.append((1, "示例玩家", "剑魂", "C", 120, None, None, None, None))
    sheet.auto_filter.ref = "A1:I2"
    return _save_workbook(workbook)


def build_roster_workbook(
    rows: list[CharacterExportRow],
    defaults: CharacterImportDefaults = DEFAULT_CHARACTER_IMPORT_DEFAULTS,
) -> bytes:
    workbook = _base_workbook(defaults)
    sheet = workbook.active
    assert sheet is not None
    for sequence, row in enumerate(rows, start=1):
        score = row.damage_score if row.role_type == "DAMAGE" else row.buffer_score
        if (
            row.role_type == "DAMAGE"
            and score is not None
            and score != score.to_integral_value()
        ):
            raise ValueError(
                f"玩家“{row.player_name}”的职业“{row.profession}”仍有小数 C 伤害，请先修改为整数"
            )
        sheet.append(
            (
                sequence,
                row.player_name,
                row.profession,
                "C" if row.role_type == "DAMAGE" else "奶",
                int(score) if row.role_type == "DAMAGE" and score is not None else score,
                _boolean_label(row.is_treasure_damage),
                _boolean_label(row.is_fixed_lead_team_buffer),
                _boolean_label(row.is_group_hunt),
                _boolean_label(row.default_raid_participant),
            )
        )
    sheet.auto_filter.ref = f"A1:I{max(1, len(rows) + 1)}"
    return _save_workbook(workbook)


def _base_workbook(defaults: CharacterImportDefaults) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "角色数据"
    sheet.append(HEADERS)
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    widths = (8, 18, 16, 10, 24, 12, 14, 12, 16)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    role_validation = DataValidation(type="list", formula1='"C,奶"')
    bool_validation = DataValidation(type="list", formula1='"是,否"')
    sheet.add_data_validation(role_validation)
    sheet.add_data_validation(bool_validation)
    role_validation.add("D2:D10001")
    bool_validation.add("F2:I10001")
    notes = workbook.create_sheet("填写说明")
    notes.append(("字段", "说明"))
    notes.append(("类型", "填写 C 或 奶"))
    notes.append(("模拟伤害亿/增益量万", "C 使用亿为单位且必须为整数，奶支持最多两位小数"))
    notes.append(
        ("是否秘宝C", f"仅 C 可填写是；空白按{_boolean_label(defaults.is_treasure_damage)}处理")
    )
    notes.append(
        (
            "玩家昵称修改",
            "导入以玩家昵称和职业匹配。需要修改玩家昵称时请先在页面修改，再导出最新人员表继续维护。",
        )
    )
    notes.append(
        (
            "固定红队奶",
            "仅奶可填写是；自动排表时固定到最高强度队伍；"
            f"空白按{_boolean_label(defaults.is_fixed_lead_team_buffer)}处理",
        )
    )
    notes.append(
        (
            "是否群猎",
            "仅 C 可填写是；当前作为角色标记保存；"
            f"空白按{_boolean_label(defaults.is_group_hunt)}处理",
        )
    )
    notes.append(
        (
            "是否参与团本",
            "控制新排表是否默认勾选；"
            f"空白按{_boolean_label(defaults.default_raid_participant)}处理；"
            "填否的启用角色仍可在候选池中手动选择",
        )
    )
    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 88
    for cell in notes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    for row in notes.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    return workbook


def _save_workbook(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def parse_character_workbook(
    content: bytes,
    max_rows: int,
    defaults: CharacterImportDefaults = DEFAULT_CHARACTER_IMPORT_DEFAULTS,
) -> list[ParsedRow]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("无法读取 Excel 文件") from exc
    sheet, field_indexes = _find_import_sheet(workbook)
    rows: list[ParsedRow] = []
    seen: set[tuple[str, str]] = set()
    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        mapped_values = {
            field: values[index] if index < len(values) else None
            for field, index in field_indexes.items()
        }
        if not any(
            value is not None and str(value).strip()
            for field, value in mapped_values.items()
            if field != "sequence"
        ):
            continue
        if len(rows) >= max_rows:
            raise ValueError(f"导入行数不能超过 {max_rows}")
        payload, errors = _parse_row(mapped_values, defaults)
        key = (payload.get("player_key", ""), payload.get("profession_key", ""))
        if all(key):
            if key in seen:
                errors.append({"code": "DUPLICATE_ROW", "message": "文件内玩家与职业重复"})
            seen.add(key)
        rows.append(ParsedRow(row_no=row_no, payload=payload, errors=errors))
    if not rows:
        raise ValueError("导入文件中没有角色数据")
    return rows


def _find_import_sheet(workbook: OpenpyxlWorkbook) -> tuple[Worksheet, dict[str, int]]:
    for sheet in workbook.worksheets:
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        indexes = {header: index for index, header in enumerate(headers) if header}
        field_indexes = {
            field: indexes[alias]
            for field, aliases in HEADER_ALIASES.items()
            for alias in aliases
            if alias in indexes
        }
        if all(field in field_indexes for field in REQUIRED_FIELDS):
            return sheet, field_indexes
    raise ValueError(
        "未找到角色数据工作表，至少需要这些列："
        "玩家昵称（或玩家称呼） | 职业 | 类型 | 模拟伤害亿/增益量万（或伤害/增益量）"
    )


def _parse_row(
    values: dict[str, Any], defaults: CharacterImportDefaults
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    player_name = _text(values.get("player_name"))
    profession = _text(values.get("profession"))
    role_raw = _text(values.get("role_type")).upper()
    errors: list[dict[str, str]] = []
    role_type = {"C": "DAMAGE", "DAMAGE": "DAMAGE", "奶": "BUFFER", "BUFFER": "BUFFER"}.get(
        role_raw
    )
    for field, value in (
        ("玩家称呼", player_name),
        ("职业", profession),
    ):
        if not value:
            errors.append({"code": "REQUIRED", "message": f"{field}不能为空"})
    if role_type is None:
        errors.append({"code": "INVALID_ROLE", "message": "类型必须为 C 或 奶"})
    score = _decimal(values.get("score"), errors)
    if role_type == "DAMAGE" and score is not None and score != score.to_integral_value():
        errors.append({"code": "DAMAGE_SCORE_NOT_INTEGER", "message": "C 伤害必须为整数"})
    treasure = _boolean(
        values.get("is_treasure_damage"),
        "是否秘宝C",
        errors,
        default=defaults.is_treasure_damage,
    )
    fixed_lead_buffer = _boolean(
        values.get("is_fixed_lead_team_buffer"),
        "固定红队奶",
        errors,
        default=defaults.is_fixed_lead_team_buffer,
    )
    group_hunt = _boolean(
        values.get("is_group_hunt"),
        "是否群猎",
        errors,
        default=defaults.is_group_hunt,
    )
    default_participant = _boolean(
        values.get("default_raid_participant"),
        "是否参与团本",
        errors,
        default=defaults.default_raid_participant,
    )
    payload: dict[str, Any] = {
        "player_name": player_name,
        "player_key": normalize_key(player_name),
        "profession": profession,
        "profession_key": normalize_key(profession),
        "role_type": role_type,
        "damage_score": str(score) if score is not None and role_type == "DAMAGE" else None,
        "buffer_score": str(score) if score is not None and role_type == "BUFFER" else None,
        "is_treasure_damage": treasure,
        "is_fixed_lead_team_buffer": fixed_lead_buffer,
        "is_group_hunt": group_hunt,
        "default_raid_participant": default_participant,
        "note": _text(values.get("note")) or None,
        "is_active": True,
        "provided_fields": sorted(values),
    }
    if not errors:
        assert role_type is not None
        try:
            CharacterCreate(
                profession=profession,
                role_type=CharacterRole(role_type),
                damage_score=payload["damage_score"],
                buffer_score=payload["buffer_score"],
                is_treasure_damage=treasure,
                is_fixed_lead_team_buffer=fixed_lead_buffer,
                is_group_hunt=group_hunt,
                default_raid_participant=default_participant,
                note=payload["note"],
            )
        except ValueError as exc:
            errors.append({"code": "INVALID_CHARACTER", "message": str(exc)})
    return payload, errors


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _decimal(value: Any, errors: list[dict[str, str]]) -> Decimal | None:
    raw = _text(value).removesuffix("亿")
    try:
        result = Decimal(raw)
    except InvalidOperation:
        errors.append({"code": "INVALID_SCORE", "message": "伤害/增益量必须是数字"})
        return None
    if result < 0:
        errors.append({"code": "INVALID_SCORE", "message": "伤害/增益量不能小于 0"})
    return result


def _boolean(
    value: Any, field: str, errors: list[dict[str, str]], *, default: bool
) -> bool:
    normalized = _text(value).casefold()
    if not normalized:
        return default
    if normalized in {"是", "y", "yes", "1", "true"}:
        return True
    if normalized in {"否", "n", "no", "0", "false"}:
        return False
    errors.append({"code": "INVALID_BOOLEAN", "message": f"{field}必须为是或否"})
    return default


def _boolean_label(value: bool) -> str:
    return "是" if value else "否"


def build_error_workbook(rows: list[tuple[int, dict[str, Any], list[dict[str, Any]]]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "错误明细"
    sheet.append(("原行号", *HEADERS, "备注", "错误"))
    for row_no, payload, errors in rows:
        role = "C" if payload.get("role_type") == "DAMAGE" else "奶"
        score = payload.get("damage_score") or payload.get("buffer_score")
        sheet.append(
            (
                row_no,
                "",
                payload.get("player_name"),
                payload.get("profession"),
                role,
                score,
                "是" if payload.get("is_treasure_damage") else "否",
                "是" if payload.get("is_fixed_lead_team_buffer") else "否",
                "是" if payload.get("is_group_hunt") else "否",
                "是" if payload.get("default_raid_participant") else "否",
                payload.get("note"),
                "；".join(str(error.get("message", "")) for error in errors),
            )
        )
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
