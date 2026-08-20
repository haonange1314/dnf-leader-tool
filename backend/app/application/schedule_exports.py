from __future__ import annotations

import io
import json
import math
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageColor, ImageDraw, ImageFont

FONT_CANDIDATES = (
    Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
    Path("/System/Library/Fonts/PingFang.ttc"),
    Path("/System/Library/Fonts/STHeiti Light.ttc"),
)


def snapshot_text(snapshot: dict[str, Any], label: str, *, draft: bool = False) -> str:
    title = f"{snapshot.get('name', '排表')} · {label}"
    lines = [f"【草稿】{title}" if draft else title]
    participants = {str(item["id"]): item for item in snapshot.get("participants", [])}
    for wave in snapshot.get("waves", []):
        lines.append("")
        lines.append(f"第 {wave['waveNo']} 波")
        cores = {str(item["participantId"]) for item in wave.get("specialAssignments", [])}
        for team in wave.get("teams", []):
            members: list[str] = []
            for slot in team.get("slots", []):
                participant = participants.get(str(slot.get("participantId")))
                if participant is None:
                    members.append("待补")
                else:
                    core = "【核心】" if str(participant["id"]) in cores else ""
                    members.append(
                        f"{participant['playerNameSnapshot']}·"
                        f"{participant['characterNameSnapshot']}{core}"
                    )
            lines.append(f"{team['displayNameSnapshot']}：{' / '.join(members)}")
    return "\n".join(lines)


def snapshot_workbook(snapshot: dict[str, Any], label: str, *, draft: bool = False) -> io.BytesIO:
    workbook = Workbook()
    overview = cast(Worksheet, workbook.active)
    overview.title = "排表总览"
    participants = {str(item["id"]): item for item in snapshot.get("participants", [])}
    assigned_ids: set[str] = set()

    _sheet_heading(overview, snapshot, label, draft=draft, column_count=9)
    overview.append(["波次", "队伍", "位置", "玩家", "角色", "职业", "类型", "评分", "核心"])
    for wave in snapshot.get("waves", []):
        cores = {str(item["participantId"]) for item in wave.get("specialAssignments", [])}
        for team in wave.get("teams", []):
            for slot in team.get("slots", []):
                participant_id = str(slot.get("participantId"))
                participant = participants.get(participant_id, {})
                if participant:
                    assigned_ids.add(participant_id)
                score = participant.get("damageScoreSnapshot") or participant.get(
                    "bufferScoreSnapshot"
                )
                overview.append(
                    [
                        wave["waveNo"],
                        team["displayNameSnapshot"],
                        slot["slotNo"],
                        participant.get("playerNameSnapshot", "待补"),
                        participant.get("characterNameSnapshot", ""),
                        participant.get("professionSnapshot", ""),
                        participant.get("roleTypeSnapshot", ""),
                        score or "",
                        "是" if participant_id in cores else "",
                    ]
                )
    _format_table(overview, 9)

    unassigned = workbook.create_sheet("未分配")
    _sheet_heading(unassigned, snapshot, label, draft=draft, column_count=6)
    unassigned.append(["玩家", "角色", "职业", "类型", "评分", "原因"])
    for participant in snapshot.get("participants", []):
        participant_id = str(participant["id"])
        if not participant.get("isSelected") or participant_id in assigned_ids:
            continue
        score = participant.get("damageScoreSnapshot") or participant.get("bufferScoreSnapshot")
        reason = participant.get("unassignedReason")
        unassigned.append(
            [
                participant.get("playerNameSnapshot", ""),
                participant.get("characterNameSnapshot", ""),
                participant.get("professionSnapshot", ""),
                participant.get("roleTypeSnapshot", ""),
                score or "",
                json.dumps(reason, ensure_ascii=False) if reason else "未说明",
            ]
        )
    _format_table(unassigned, 6)

    strengths = workbook.create_sheet("强度统计")
    _sheet_heading(strengths, snapshot, label, draft=draft, column_count=6)
    strengths.append(["波次", "队伍", "组成", "C 强度", "奶强度", "波次总强度"])
    for wave in snapshot.get("waves", []):
        for team in wave.get("teams", []):
            strengths.append(
                [
                    wave["waveNo"],
                    team["displayNameSnapshot"],
                    team.get("compositionCode", ""),
                    team.get("damageTotal", 0),
                    team.get("bufferTotal", 0),
                    f"C {wave.get('damageTotal', 0)} / 奶 {wave.get('bufferTotal', 0)}",
                ]
            )
    _format_table(strengths, 6)

    issues = workbook.create_sheet("问题清单")
    _sheet_heading(issues, snapshot, label, draft=draft, column_count=3)
    issues.append(["级别", "代码", "详情"])
    for issue in snapshot.get("issues", []):
        issues.append(
            [
                issue.get("severity", ""),
                issue.get("code", ""),
                json.dumps(issue.get("message_params", {}), ensure_ascii=False),
            ]
        )
    _format_table(issues, 3)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def snapshot_png(snapshot: dict[str, Any], label: str, *, draft: bool = False) -> io.BytesIO:
    waves = snapshot.get("waves", [])
    team_count = max((len(wave.get("teams", [])) for wave in waves), default=1)
    width = max(1200, min(3200, 96 + team_count * 360))
    horizontal_margin = 48
    team_gap = 16
    team_width = max(
        280,
        (width - horizontal_margin * 2 - team_gap * (team_count - 1)) // team_count,
    )
    member_columns = max(1, team_width // 280)
    wave_heights = [_wave_image_height(wave, member_columns) for wave in waves]
    height = 150 + sum(wave_heights) + max(0, len(waves) - 1) * 20 + 48
    image = Image.new("RGBA", (width, max(height, 240)), "#f2f0eb")
    draw = ImageDraw.Draw(image)
    title_font = _font(36)
    heading_font = _font(24)
    body_font = _font(20)
    small_font = _font(17)
    draw.text(
        (horizontal_margin, 34),
        str(snapshot.get("name", "排表")),
        font=title_font,
        fill="#202124",
    )
    draw.text((horizontal_margin, 88), label, font=body_font, fill="#6b7280")
    if draft:
        draw.rounded_rectangle((width - 180, 34, width - 48, 88), radius=12, fill="#d44a3a")
        draw.text((width - 147, 45), "草稿", font=heading_font, fill="#ffffff")

    participants = {str(item["id"]): item for item in snapshot.get("participants", [])}
    y = 130
    for wave, wave_height in zip(waves, wave_heights, strict=True):
        draw.rounded_rectangle(
            (horizontal_margin, y, width - horizontal_margin, y + wave_height),
            radius=18,
            fill="#ffffff",
            outline="#ded9cf",
            width=2,
        )
        draw.text(
            (horizontal_margin + 24, y + 18),
            f"第 {wave['waveNo']} 波",
            font=heading_font,
            fill="#202124",
        )
        stats = f"C {wave.get('damageTotal', 0)} 亿  ·  奶 {wave.get('bufferTotal', 0)}"
        stats_width = draw.textlength(stats, font=body_font)
        draw.text(
            (width - horizontal_margin - 24 - stats_width, y + 21),
            stats,
            font=body_font,
            fill="#6b7280",
        )
        teams = wave.get("teams", [])
        current_team_width = (
            width - horizontal_margin * 2 - 48 - team_gap * max(0, len(teams) - 1)
        ) // max(1, len(teams))
        card_y = y + 66
        cores = {str(item["participantId"]) for item in wave.get("specialAssignments", [])}
        for index, team in enumerate(teams):
            card_x = horizontal_margin + 24 + index * (current_team_width + team_gap)
            card_right = card_x + current_team_width
            team_color = _safe_color(str(team.get("displayColorSnapshot", "#6b7280")))
            draw.rounded_rectangle(
                (card_x, card_y, card_right, y + wave_height - 20),
                radius=12,
                fill="#faf9f6",
                outline="#ece8df",
            )
            draw.rounded_rectangle(
                (card_x, card_y, card_right, card_y + 8),
                radius=4,
                fill=team_color,
            )
            team_title = (
                f"{team.get('displayNameSnapshot', team.get('teamKey', '队伍'))}"
                f"  ·  {team.get('compositionCode', '')}"
            )
            draw.text(
                (card_x + 14, card_y + 18),
                _fit_text(draw, team_title, heading_font, current_team_width - 28),
                font=heading_font,
                fill="#292724",
            )
            slots = team.get("slots", [])
            columns = max(1, current_team_width // 280)
            column_width = (current_team_width - 28) // columns
            for slot_index, slot in enumerate(slots):
                row = slot_index // columns
                column = slot_index % columns
                participant = participants.get(str(slot.get("participantId")))
                member_y = card_y + 62 + row * 42
                member_x = card_x + 14 + column * column_width
                if participant is None:
                    member_text = f"{slot.get('slotNo', slot_index + 1)}. 待补"
                    color = "#9ca3af"
                else:
                    role = "C" if participant.get("roleTypeSnapshot") == "DAMAGE" else "奶"
                    core = " · 核心" if str(participant["id"]) in cores else ""
                    member_text = (
                        f"[{role}] {participant.get('playerNameSnapshot', '')} · "
                        f"{participant.get('characterNameSnapshot', '')}{core}"
                    )
                    color = "#292724"
                draw.text(
                    (member_x, member_y),
                    _fit_text(draw, member_text, small_font, column_width - 10),
                    font=small_font,
                    fill=color,
                )
        if draft:
            watermark = "草稿 · 非最终版本"
            watermark_width = draw.textlength(watermark, font=title_font)
            draw.text(
                ((width - watermark_width) / 2, y + wave_height / 2 - 20),
                watermark,
                font=title_font,
                fill="#f2d8d4",
            )
        y += wave_height + 20

    output = io.BytesIO()
    image.convert("RGB").save(output, format="PNG", optimize=True)
    output.seek(0)
    return output


def _wave_image_height(wave: dict[str, Any], member_columns: int) -> int:
    max_slots = max((len(team.get("slots", [])) for team in wave.get("teams", [])), default=1)
    return 154 + math.ceil(max_slots / max(1, member_columns)) * 42


def _font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in FONT_CANDIDATES:
        if path.exists():
            return ImageFont.truetype(str(path), size=size)
    return ImageFont.load_default(size=size)


def _safe_color(value: str) -> str:
    try:
        ImageColor.getrgb(value)
    except ValueError:
        return "#6b7280"
    return value


def _fit_text(
    draw: ImageDraw.ImageDraw,
    value: str,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    maximum_width: int,
) -> str:
    if draw.textlength(value, font=font) <= maximum_width:
        return value
    suffix = "…"
    trimmed = value
    while trimmed and draw.textlength(f"{trimmed}{suffix}", font=font) > maximum_width:
        trimmed = trimmed[:-1]
    return f"{trimmed}{suffix}"


def _sheet_heading(
    sheet: Worksheet,
    snapshot: dict[str, Any],
    label: str,
    *,
    draft: bool,
    column_count: int,
) -> None:
    if draft:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
        cell = sheet.cell(row=1, column=1, value="草稿 · 非最终发布版本")
        cell.fill = PatternFill("solid", fgColor="D44A3A")
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
    sheet.append([snapshot.get("name", "排表"), label])


def _format_table(sheet: Worksheet, column_count: int) -> None:
    for column in range(1, column_count + 1):
        sheet.column_dimensions[chr(64 + column)].width = 20
    sheet.freeze_panes = f"A{4 if sheet['A1'].value == '草稿 · 非最终发布版本' else 3}"
